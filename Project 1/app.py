import openpyxl as xl

wb=xl.load_workbook("transactions.xlsx")

sheet=wb["Sheet1"]

sheet.cell(1,4).value="Updated Value"

for i in range(2,sheet.max_row+1):
    newVal=sheet.cell(i,3).value*100
    sheet.cell(i,4).value=newVal

wb.save("transaction3.xlsx")